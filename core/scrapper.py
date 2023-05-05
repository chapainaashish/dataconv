import os

import openpyxl
import pandas as pd


class DataScrap:
    """Scraps Data from csv and excel file"""

    def __init__(self, file):
        self.file = f"media/uploads/{file}"

    def filetype(self):
        """Determines the file type"""
        extension = os.path.splitext(self.file)[1]
        return extension

    def read_csv(self):
        """Read csv file"""
        data = pd.read_csv(self.file).filter(regex="(?i)email")
        return data

    def remove_empty_cells(self, sheet):
        """Remove empty cells from excel file"""
        # iterate the sheet by rows
        for row in sheet.iter_rows():
            # all() return False if all of the row value is None
            if not all(cell.value for cell in row):
                # delete the empty row
                sheet.delete_rows(row[0].row, 1)
                # recursively call the remove_empty_cells() with modified sheet data
                self.remove_empty_cells(sheet)

    def normalize_excel(self):
        """Normalize the excel file to read it"""
        book = openpyxl.load_workbook(self.file)
        for sheet in book.sheetnames:
            active = book[sheet]
            for row in active:
                self.remove_empty_cells(active)
        book.save(self.file)

    def read_excel(self):
        """Read the excel file"""
        data = pd.read_excel(self.file).filter(regex="(?i)email")
        self.normalize_excel()
        return data

    def get_data(self):
        """Get the data from raw files"""
        extension = self.filetype()
        if extension == ".csv":
            data = self.read_csv()
        elif extension == ".xlsx":
            data = self.read_excel()
        return data
